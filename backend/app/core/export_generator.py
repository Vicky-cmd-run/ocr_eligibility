"""
Export Generator — CSV and Excel export for batch results.
"""
import io
import logging
from typing import List, Dict, Any

import pandas as pd

logger = logging.getLogger(__name__)


def _build_row(doc_data: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten a document result dict into a flat export row."""
    candidate = doc_data.get("candidate", {}) or {}
    extraction = doc_data.get("extraction_result", {}) or {}
    eligibility = doc_data.get("eligibility_result", {}) or {}
    doc = doc_data.get("document", {}) or {}

    rejection_reasons = eligibility.get("rejection_reasons") or []
    if isinstance(rejection_reasons, list):
        rejection_str = "; ".join(rejection_reasons)
    else:
        rejection_str = str(rejection_reasons)

    return {
        "Document ID": str(doc.get("id", "")),
        "Filename": doc.get("original_filename", ""),
        "Status": doc.get("status", ""),
        "Candidate Name": candidate.get("name", ""),
        "Register Number": candidate.get("register_number", ""),
        "Physics %": _fmt(extraction.get("physics_percentage")),
        "Chemistry %": _fmt(extraction.get("chemistry_percentage")),
        "Mathematics %": _fmt(extraction.get("mathematics_percentage")),
        "PCM Cutoff": _fmt(extraction.get("pcm_cutoff")),
        "Cutoff Formula": extraction.get("cutoff_formula_used", ""),
        "Overall %": _fmt(extraction.get("overall_percentage")),
        "OCR Confidence": _fmt(extraction.get("combined_confidence")),
        "Eligibility Status": eligibility.get("status", ""),
        "Physics Passed": eligibility.get("physics_passed", ""),
        "Chemistry Passed": eligibility.get("chemistry_passed", ""),
        "Mathematics Passed": eligibility.get("mathematics_passed", ""),
        "Overall Passed": eligibility.get("overall_passed", ""),
        "Rejection Reasons": rejection_str,
        "Manually Reviewed": eligibility.get("is_manually_reviewed", False),
        "Review Notes": eligibility.get("review_notes", ""),
    }


def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


def generate_csv(rows: List[Dict[str, Any]]) -> bytes:
    """Generate a CSV export from a list of document result dicts."""
    flat_rows = [_build_row(r) for r in rows]
    df = pd.DataFrame(flat_rows)
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    return buf.getvalue()


def generate_excel(rows: List[Dict[str, Any]]) -> bytes:
    """Generate an Excel (.xlsx) export from document result dicts."""
    flat_rows = [_build_row(r) for r in rows]
    df = pd.DataFrame(flat_rows)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")

        # Auto-size columns
        worksheet = writer.sheets["Results"]
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Color-code eligibility column
        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        status_col_idx = list(df.columns).index("Eligibility Status") + 1
        for row_idx, row_data in enumerate(flat_rows, start=2):
            cell = worksheet.cell(row=row_idx, column=status_col_idx)
            if row_data.get("Eligibility Status") == "ELIGIBLE":
                cell.fill = green_fill
            elif row_data.get("Eligibility Status") == "NOT_ELIGIBLE":
                cell.fill = red_fill
            elif row_data.get("Eligibility Status") == "REVIEW_REQUIRED":
                cell.fill = orange_fill

    buf.seek(0)
    return buf.getvalue()
